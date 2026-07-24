from src.database.mongodb_connection import MongoDB
from datetime import datetime, timedelta
from bson import ObjectId
import uuid
import csv
import io

ROLE_HIERARCHY = {
    'super_admin': 4,
    'finance_admin': 3,
    'finance_operator': 2,
    'readonly': 1,
}

class FinanceAPI:
    def __init__(self):
        self.db = MongoDB()

    def _resolve_id(self, id_val):
        if isinstance(id_val, ObjectId):
            return id_val
        try:
            return ObjectId(id_val)
        except Exception:
            return id_val

    def _check_permission(self, user, required_role):
        role = user.get('role', 'readonly')
        return ROLE_HIERARCHY.get(role, 0) >= ROLE_HIERARCHY.get(required_role, 0)

    def get_user_role(self, user):
        return user.get('role', 'readonly')

    def update_user_role(self, user_id, new_role, admin_user):
        if not self._check_permission(admin_user, 'super_admin'):
            return {'status': 'error', 'message': 'Only Super Admin can change roles'}
        if new_role not in ROLE_HIERARCHY:
            return {'status': 'error', 'message': f'Invalid role. Valid roles: {", ".join(ROLE_HIERARCHY.keys())}'}
        self.db.users.update_one(
            {'_id': self._resolve_id(user_id)},
            {'$set': {'role': new_role, 'updated_at': datetime.now()}}
        )
        return {'status': 'success', 'message': f'User role updated to {new_role}'}

    def get_all_roles(self):
        return {
            'roles': [
                {'id': 'super_admin', 'name': 'Super Admin', 'level': 4, 'description': 'Full access to all modules'},
                {'id': 'finance_admin', 'name': 'Finance Admin', 'level': 3, 'description': 'Manage finance module, bank accounts, and transactions'},
                {'id': 'finance_operator', 'name': 'Finance Operator', 'level': 2, 'description': 'Add and view transactions, limited access'},
                {'id': 'readonly', 'name': 'Read-only', 'level': 1, 'description': 'View-only access to finance data'},
            ]
        }

    # ==================== BANK ACCOUNTS ====================

    def get_bank_accounts(self, user):
        accounts = []
        for acc in self.db.bank_accounts.find().sort('created_at', -1):
            balance = self._calculate_balance(str(acc['_id']))
            accounts.append({
                'id': str(acc['_id']),
                'bank_name': acc.get('bank_name', ''),
                'account_name': acc.get('account_name', ''),
                'account_number': acc.get('account_number', ''),
                'account_type': acc.get('account_type', 'savings'),
                'initial_balance': acc.get('initial_balance', 0),
                'current_balance': balance,
                'currency': acc.get('currency', 'BDT'),
                'status': acc.get('status', 'active'),
                'notes': acc.get('notes', ''),
                'created_at': acc['created_at'].strftime('%Y-%m-%d %H:%M:%S') if acc.get('created_at') else '',
                'updated_at': acc['updated_at'].strftime('%Y-%m-%d %H:%M:%S') if acc.get('updated_at') else '',
            })
        return accounts

    def add_bank_account(self, data, user):
        if not self._check_permission(user, 'finance_admin'):
            return {'status': 'error', 'message': 'Insufficient permissions. Finance Admin or higher required.'}
        account = {
            'bank_name': data.get('bank_name', ''),
            'account_name': data.get('account_name', ''),
            'account_number': data.get('account_number', ''),
            'account_type': data.get('account_type', 'savings'),
            'initial_balance': float(data.get('initial_balance', 0)),
            'currency': data.get('currency', 'BDT'),
            'status': 'active',
            'notes': data.get('notes', ''),
            'created_by': user.get('email', ''),
            'created_at': datetime.now(),
            'updated_at': datetime.now(),
        }
        result = self.db.bank_accounts.insert_one(account)
        return {'status': 'success', 'id': str(result.inserted_id), 'message': 'Bank account added successfully'}

    def update_bank_account(self, account_id, data, user):
        if not self._check_permission(user, 'finance_admin'):
            return {'status': 'error', 'message': 'Insufficient permissions. Finance Admin or higher required.'}
        allowed_fields = ['bank_name', 'account_name', 'account_number', 'account_type', 'initial_balance', 'currency', 'status', 'notes']
        updates = {k: v for k, v in data.items() if k in allowed_fields}
        if 'initial_balance' in updates:
            updates['initial_balance'] = float(updates['initial_balance'])
        updates['updated_at'] = datetime.now()
        self.db.bank_accounts.update_one(
            {'_id': self._resolve_id(account_id)},
            {'$set': updates}
        )
        return {'status': 'success', 'message': 'Bank account updated successfully'}

    def delete_bank_account(self, account_id, user):
        if not self._check_permission(user, 'super_admin'):
            return {'status': 'error', 'message': 'Only Super Admin can delete bank accounts'}
        self.db.bank_accounts.delete_one({'_id': self._resolve_id(account_id)})
        self.db.finance_transactions.delete_many({'bank_account_id': account_id})
        return {'status': 'success', 'message': 'Bank account and related transactions deleted'}

    def _calculate_balance(self, account_id):
        account = self.db.bank_accounts.find_one({'_id': self._resolve_id(account_id)})
        if not account:
            return 0
        initial = float(account.get('initial_balance', 0))
        income = 0
        expense = 0
        for t in self.db.finance_transactions.find({'bank_account_id': account_id}):
            amount = float(t.get('amount', 0))
            if t.get('type') == 'income':
                income += amount
            elif t.get('type') == 'expense':
                expense += amount
        return initial + income - expense

    # ==================== TRANSACTIONS ====================

    def get_transactions(self, user, params=None):
        params = params or {}
        page = int(params.get('page', 1))
        per_page = int(params.get('per_page', 50))
        account_id = params.get('bank_account_id', '')
        tx_type = params.get('type', '')
        category = params.get('category', '')
        subcategory = params.get('subcategory', '')
        date_from = params.get('date_from', '')
        date_to = params.get('date_to', '')
        search = params.get('search', '')

        query = {}
        if account_id:
            query['bank_account_id'] = account_id
        if tx_type:
            query['type'] = tx_type
        if category:
            query['category'] = category
        if subcategory:
            query['subcategory'] = subcategory
        if date_from or date_to:
            query['date'] = {}
            if date_from:
                try:
                    query['date']['$gte'] = datetime.strptime(date_from, '%Y-%m-%d')
                except:
                    pass
            if date_to:
                try:
                    query['date']['$lte'] = datetime.strptime(date_to + ' 23:59:59', '%Y-%m-%d %H:%M:%S')
                except:
                    pass
        if search:
            query['$or'] = [
                {'description': {'$regex': search, '$options': 'i'}},
                {'category': {'$regex': search, '$options': 'i'}},
                {'reference': {'$regex': search, '$options': 'i'}},
            ]

        total = self.db.finance_transactions.count_documents(query)
        transactions = []
        for t in list(self.db.finance_transactions.find(query).sort('date', -1).skip((page - 1) * per_page).limit(per_page)):
            transactions.append({
                'id': str(t['_id']),
                'bank_account_id': t.get('bank_account_id', ''),
                'type': t.get('type', ''),
                'amount': t.get('amount', 0),
                'category': t.get('category', ''),
                'subcategory': t.get('subcategory', ''),
                'description': t.get('description', ''),
                'reference': t.get('reference', ''),
                'date': t['date'].strftime('%Y-%m-%d') if t.get('date') else '',
                'created_by': t.get('created_by', ''),
                'created_at': t['created_at'].strftime('%Y-%m-%d %H:%M:%S') if t.get('created_at') else '',
            })

        return {
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': max(1, (total + per_page - 1) // per_page),
            'transactions': transactions,
        }

    def add_transaction(self, data, user):
        if not self._check_permission(user, 'finance_operator'):
            return {'status': 'error', 'message': 'Insufficient permissions. Finance Operator or higher required.'}
        tx = {
            'bank_account_id': data.get('bank_account_id', ''),
            'type': data.get('type', 'expense'),
            'amount': float(data.get('amount', 0)),
            'category': data.get('category', ''),
            'subcategory': data.get('subcategory', ''),
            'description': data.get('description', ''),
            'reference': data.get('reference', ''),
            'date': datetime.strptime(data.get('date', datetime.now().strftime('%Y-%m-%d')), '%Y-%m-%d') if data.get('date') else datetime.now(),
            'created_by': user.get('name', user.get('email', '')),
            'created_at': datetime.now(),
        }
        if tx['amount'] <= 0:
            return {'status': 'error', 'message': 'Amount must be greater than 0'}
        result = self.db.finance_transactions.insert_one(tx)
        return {'status': 'success', 'id': str(result.inserted_id), 'message': 'Transaction added successfully'}

    def update_transaction(self, tx_id, data, user):
        if not self._check_permission(user, 'finance_operator'):
            return {'status': 'error', 'message': 'Insufficient permissions. Finance Operator or higher required.'}
        allowed_fields = ['bank_account_id', 'type', 'amount', 'category', 'subcategory', 'description', 'reference', 'date']
        updates = {k: v for k, v in data.items() if k in allowed_fields}
        if 'amount' in updates:
            updates['amount'] = float(updates['amount'])
            if updates['amount'] <= 0:
                return {'status': 'error', 'message': 'Amount must be greater than 0'}
        if 'date' in updates and updates['date']:
            try:
                updates['date'] = datetime.strptime(updates['date'], '%Y-%m-%d')
            except:
                pass
        self.db.finance_transactions.update_one(
            {'_id': self._resolve_id(tx_id)},
            {'$set': updates}
        )
        return {'status': 'success', 'message': 'Transaction updated successfully'}

    def delete_transaction(self, tx_id, user):
        if not self._check_permission(user, 'finance_admin'):
            return {'status': 'error', 'message': 'Insufficient permissions. Finance Admin or higher required.'}
        self.db.finance_transactions.delete_one({'_id': self._resolve_id(tx_id)})
        return {'status': 'success', 'message': 'Transaction deleted successfully'}

    # ==================== IMPORT TRANSACTIONS ====================

    def import_transactions(self, file_content, filename, mapping, user):
        if not self._check_permission(user, 'finance_operator'):
            return {'status': 'error', 'message': 'Insufficient permissions. Finance Operator or higher required.'}

        errors = []
        imported = 0
        skipped = 0

        try:
            if filename.endswith('.csv'):
                reader = csv.DictReader(io.StringIO(file_content))
                rows = list(reader)
            elif filename.endswith(('.xls', '.xlsx')):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(file_content.encode('latin-1')))
                    ws = wb.active
                    headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {headers[i]: str(row[i]) if row[i] is not None else '' for i in range(len(headers)) if i < len(row)}
                        rows.append(row_dict)
                except ImportError:
                    return {'status': 'error', 'message': 'Excel file support requires openpyxl. Please save as CSV and try again.'}
            else:
                return {'status': 'error', 'message': 'Unsupported file format. Please use CSV, XLS, or XFX files.'}
        except Exception as e:
            return {'status': 'error', 'message': f'Error reading file: {str(e)}'}

        if not rows:
            return {'status': 'error', 'message': 'No data rows found in the file'}

        for i, row in enumerate(rows):
            try:
                tx_data = {}
                for field, col_name in mapping.items():
                    if col_name and col_name in row:
                        tx_data[field] = row[col_name].strip()

                if not tx_data.get('amount'):
                    errors.append(f'Row {i + 2}: Missing amount')
                    skipped += 1
                    continue

                try:
                    amount_str = tx_data['amount'].replace(',', '').replace(' ', '').replace('৳', '').replace('$', '').replace('BDT', '')
                    tx_data['amount'] = float(amount_str)
                except (ValueError, AttributeError):
                    errors.append(f'Row {i + 2}: Invalid amount "{tx_data.get("amount", "")}"')
                    skipped += 1
                    continue

                if tx_data.get('date'):
                    try:
                        datetime.strptime(tx_data['date'], '%Y-%m-%d')
                    except ValueError:
                        try:
                            parsed = datetime.strptime(tx_data['date'], '%d/%m/%Y')
                            tx_data['date'] = parsed.strftime('%Y-%m-%d')
                        except ValueError:
                            try:
                                parsed = datetime.strptime(tx_data['date'], '%m/%d/%Y')
                                tx_data['date'] = parsed.strftime('%Y-%m-%d')
                            except ValueError:
                                tx_data['date'] = datetime.now().strftime('%Y-%m-%d')
                else:
                    tx_data['date'] = datetime.now().strftime('%Y-%m-%d')

                tx_data.setdefault('type', 'expense')
                if tx_data['type'] not in ('income', 'expense', 'transfer'):
                    tx_data['type'] = 'expense'
                tx_data.setdefault('category', 'Uncategorized')
                tx_data.setdefault('subcategory', '')
                tx_data.setdefault('description', '')
                tx_data.setdefault('reference', '')

                self.db.finance_transactions.insert_one({
                    'bank_account_id': tx_data.get('bank_account_id', mapping.get('bank_account_id', '')),
                    'type': tx_data['type'],
                    'amount': tx_data['amount'],
                    'category': tx_data['category'],
                    'subcategory': tx_data['subcategory'],
                    'description': tx_data['description'],
                    'reference': tx_data['reference'],
                    'date': datetime.strptime(tx_data['date'], '%Y-%m-%d'),
                    'created_by': user.get('name', user.get('email', '')),
                    'created_at': datetime.now(),
                })
                imported += 1
            except Exception as e:
                errors.append(f'Row {i + 2}: {str(e)}')
                skipped += 1

        return {
            'status': 'success',
            'imported': imported,
            'skipped': skipped,
            'errors': errors[:50],
            'message': f'Imported {imported} transactions, skipped {skipped}',
        }

    # ==================== CATEGORIES ====================

    def get_categories(self):
        return {
            'income': {
                'label': 'Income',
                'subcategories': [
                    'Sales Revenue', 'Service Revenue', 'Subscription', 'Commission',
                    'Interest Income', 'Investment Return', 'Rental Income',
                    'Freelance', 'Donation', 'Grant', 'Other Income'
                ]
            },
            'expense': {
                'label': 'Expense',
                'subcategories': [
                    'Salaries', 'Rent', 'Utilities', 'Office Supplies',
                    'Marketing', 'Software', 'Hosting', 'Domain',
                    'Travel', 'Insurance', 'Tax', 'Legal Fees',
                    'Maintenance', 'Training', 'Miscellaneous'
                ]
            },
            'transfer': {
                'label': 'Transfer',
                'subcategories': [
                    'Between Accounts', 'To Savings', 'From Savings',
                    'Investment Transfer', 'Loan Payment', 'Other Transfer'
                ]
            }
        }

    # ==================== FINANCE SUMMARY ====================

    def get_finance_summary(self, user, params=None):
        params = params or {}
        month = int(params.get('month', datetime.now().month))
        year = int(params.get('year', datetime.now().year))

        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)

        query = {'date': {'$gte': start_date, '$lt': end_date}}
        total_income = 0
        total_expense = 0
        for t in self.db.finance_transactions.find(query):
            amount = float(t.get('amount', 0))
            if t.get('type') == 'income':
                total_income += amount
            elif t.get('type') == 'expense':
                total_expense += amount

        accounts = self.get_bank_accounts(user)
        total_balance = sum(a['current_balance'] for a in accounts)

        return {
            'total_income': total_income,
            'total_expense': total_expense,
            'net_balance': total_income - total_expense,
            'total_balance': total_balance,
            'accounts_count': len(accounts),
            'month': month,
            'year': year,
        }
